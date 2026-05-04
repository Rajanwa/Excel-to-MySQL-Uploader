import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="pandas")

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.pdfgen import canvas as rl_canvas
import os, sys, logging, threading, re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import mysql.connector
import datetime
from datetime import datetime as dt, timedelta
from typing import List, Optional


# ─────────────────────────────────────────────────────────────────────────────
# ConfigManager
# ─────────────────────────────────────────────────────────────────────────────
class ConfigManager:
    TABLE_FORMATS = {
        'anateck.idfc':    ["Transaction Date","Payment date","Narrative","Customer Reference No","Cheque No","Debit","Credit","Running Balance"],
        'attroidfc.idfc':  ["Transaction Date","Payment date","Narrative","Customer Reference No","Cheque No","Debit","Credit","Running Balance"],
        'anateck.yes':     ["Transaction Date","Value date","Transaction Description","Reference No","Debit Amount","Credit Amount","Running Balance"],
        'attroidfc.yes':   ["Transaction Date","Value date","Transaction Description","Reference No","Debit Amount","Credit Amount","Running Balance"],
        'anateck.kotak':   ["#","TRANSACTION DATE","VALUE DATE","TRANSACTION DETAILS","CHQ / REF NO.","DEBIT/CREDIT(₹)","BALANCE(₹)"],
        'attroidfc.kotak': ["#","TRANSACTION DATE","VALUE DATE","TRANSACTION DETAILS","CHQ / REF NO.","DEBIT/CREDIT(₹)","BALANCE(₹)"],
    }

    @classmethod
    def get_table_instructions(cls, db_name: str, table_name: str) -> str:
        instructions = {
            'fbpayripe.payin':        'Remove First Col. Created Time and Then Upload',
            'fbpayripe.payout':       'Upload as it is.',
            'fbpayripe.wallet':       'Upload as it is.',
            'anateck.payout':         'Change DATE FORMATS TO YYYY-MM-DD HH:MM:SS and Remove commas in amount col.',
            'anateck.wallet':         'Change DATE FORMATS TO YYYY-MM-DD HH:MM:SS and Remove commas in amount col.',
            'anateck.payon_payout':   'Change DATE FORMATS TO YYYY-MM-DD HH:MM:SS and Remove commas in amount col.',
            'anateck.payon_owallet':  'Change DATE FORMATS TO YYYY-MM-DD HH:MM:SS and Remove commas in amount col.',
            'anateck.idfc':           'Extract specific table format data and process UTR/UID',
            'anateck.yes':            'Extract specific table format data and process dates/amounts',
            'anateck.kotak':          'Extract specific table format data and process dates/amounts',
            'attroidfc.payout':       'Change DATE FORMATS TO YYYY-MM-DD HH:MM:SS and Remove commas in amount col.',
            'attroidfc.wallet':       'Change DATE FORMATS TO YYYY-MM-DD HH:MM:SS and Remove commas in amount col.',
            'attroidfc.payon_payout': 'Change DATE FORMATS TO YYYY-MM-DD HH:MM:SS and Remove commas in amount col.',
            'attroidfc.payon_owallet':'Change DATE FORMATS TO YYYY-MM-DD HH:MM:SS and Remove commas in amount col.',
            'attroidfc.idfc':         'Extract specific table format data and process UTR/UID',
            'attroidfc.yes':          'Extract specific table format data and process dates/amounts',
            'attroidfc.kotak':        'Extract specific table format data and process dates/amounts',
        }
        return instructions.get(
            f"{db_name}.{table_name}",
            "Upload as it is. Empty columns will be set to NULL. Dates should be in YYYY-MM-DD HH:MM:SS format."
        )

    @classmethod
    def get_expected_columns(cls, db_name: str, table_name: str) -> Optional[List[str]]:
        return cls.TABLE_FORMATS.get(f"{db_name}.{table_name}")

    @classmethod
    def requires_table_format_extraction(cls, db_name: str, table_name: str) -> bool:
        return f"{db_name}.{table_name}" in cls.TABLE_FORMATS


# ─────────────────────────────────────────────────────────────────────────────
# LoggerManager
# ─────────────────────────────────────────────────────────────────────────────
class LoggerManager:
    @staticmethod
    def setup_logger(log_file='data_uploader.log'):
        log_dir = os.path.join(os.getcwd(), 'logs')
        os.makedirs(log_dir, exist_ok=True)
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s | %(levelname)s | %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S',
            handlers=[
                logging.FileHandler(os.path.join(log_dir, log_file), encoding='utf-8'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        return logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# TableFormatExtractor
# ─────────────────────────────────────────────────────────────────────────────
class TableFormatExtractor:
    @staticmethod
    def extract_table_data(df, expected_columns, logger):
        header_row_idx = TableFormatExtractor._find_header_row(df, expected_columns, logger)
        if header_row_idx is None:
            raise ValueError(f"Could not find table with expected columns: {expected_columns}")
        table_df = df.iloc[header_row_idx:].copy()
        table_df.columns = table_df.iloc[0]
        table_df = table_df.iloc[1:]
        available = [c for c in expected_columns if c in table_df.columns]
        table_df = table_df[available].dropna(how='all').reset_index(drop=True)
        return table_df

    @staticmethod
    def _find_header_row(df, expected_columns, logger):
        norm = lambda t: "" if pd.isna(t) else str(t).strip().lower()
        ne = [norm(c) for c in expected_columns]
        for idx in range(min(50, len(df))):
            row = df.iloc[idx].values
            nr = [norm(v) for v in row]
            ratio = sum(1 for ec in ne if ec in nr) / len(ne)
            if ratio >= 0.7:
                non_empty = sum(1 for v in row if pd.notna(v) and str(v).strip())
                if non_empty >= len(expected_columns) * 0.5:
                    return idx
        return None


# ─────────────────────────────────────────────────────────────────────────────
# DataProcessor
# ─────────────────────────────────────────────────────────────────────────────
class DataProcessor:
    @staticmethod
    def process_data(df, db_name, table_name, logger):
        processed_df = df.copy()
        processed_df = DataProcessor.protect_scientific_notation(processed_df)
        processed_df = DataProcessor.format_dates(processed_df, logger)
        processed_df = DataProcessor.clean_amount_columns(processed_df)
        if ConfigManager.requires_table_format_extraction(db_name, table_name):
            ec = ConfigManager.get_expected_columns(db_name, table_name)
            if ec:
                processed_df = TableFormatExtractor.extract_table_data(processed_df, ec, logger)
        processed_df = DataProcessor.apply_table_specific_processing(processed_df, db_name, table_name, logger)
        return processed_df

    @staticmethod
    def apply_table_specific_processing(df, db_name, table_name, logger):
        if table_name.lower() in ['idfc', 'yes', 'kotak']:
            df = DataProcessor.process_bank_statement_columns(df, table_name.lower(), logger)
        if "Remove First Col. Created Time" in ConfigManager.get_table_instructions(db_name, table_name):
            df = df.iloc[:, 1:] if len(df.columns) > 0 else df
        return df

    @staticmethod
    def process_bank_statement_columns(df, bank_name, logger):
        for col in [c for c in df.columns if 'date' in c.lower()]:
            df[col] = DataProcessor._convert_date_column(df[col], logger)
        for col in [c for c in df.columns if any(k in c.lower() for k in ['amount','debit','credit','balance','value'])]:
            df[col] = DataProcessor._clean_numeric_value(df[col], logger)
        if bank_name == 'idfc':
            df = DataProcessor.process_idfc_reference_numbers(df, logger)
        return df

    @staticmethod
    def process_idfc_reference_numbers(df, logger):
        if 'Narrative' not in df.columns:
            return df
        p = re.compile(r'UPI/MOB/(\d{10,15})/([A-Za-z0-9]{8,30})')
        for idx, row in df.iterrows():
            narrative = str(row['Narrative']).strip()
            m = p.search(narrative)
            if m:
                df.at[idx, 'Customer Reference No'] = m.group(1)
                df.at[idx, 'Cheque No'] = m.group(2)
        return df

    @staticmethod
    def _clean_numeric_value(series, logger):
        cleaned = series.astype(str).str.replace(r'[₹,$€£,\s]', '', regex=True)
        cleaned = cleaned.str.replace(r'\(([\d.]+)\)', r'-\1', regex=True)
        return pd.to_numeric(cleaned, errors='coerce').fillna(0)

    @staticmethod
    def protect_scientific_notation(df):
        for col in df.columns:
            if any(k in str(col).lower() for k in ['date','time']):
                continue
            if df[col].dtype == 'object':
                mask = df[col].astype(str).str.contains(r'^\d+\.?\d*[Ee][+-]?\d+$', na=False)
                if mask.any():
                    df.loc[mask, col] = df.loc[mask, col].apply(
                        lambda x: format(float(x), '.{}f'.format(len(str(x).split('.')[1]) if '.' in str(x) else 0))
                    )
        return df

    @staticmethod
    def format_dates(df, logger):
        date_kw = ['date','time','datetime','timestamp','created','modified','processed','reversed','scheduled']
        for col in df.columns:
            if not any(k in str(col).lower() for k in date_kw):
                continue
            original = df[col].copy()
            s = df[col].astype(str).str.strip().replace(['','NA','N/A','nan','None','NULL','NaT'], None)
            result = s.copy()
            bad = s.str.contains(r'1900-01-00', na=False)
            result[bad] = '1900-01-01 00:00:00'
            handlers = [
                {'name':'period-sep', 'pattern':r'\d{2}-\d{2}-\d{4} \d{2}\.\d{2}\.\d{2}',
                 'preprocessor': lambda x: re.sub(r'(\d{2})-(\d{2})-(\d{4}) (\d{2})\.(\d{2})\.(\d{2})',r'\3-\2-\1 \4:\5:\6',x),
                 'format':'%Y-%m-%d %H:%M:%S'},
                {'name':'AMPM',    'pattern':r'\d{1,2}/\d{1,2}/\d{2,4} \d{1,2}:\d{2}:\d{2} [AP]M', 'format':'%d/%m/%Y %I:%M:%S %p'},
                {'name':'ISO8601', 'pattern':r'\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}', 'format':'ISO8601'},
                {'name':'YYYY-MM-DD','pattern':r'^\d{4}-\d{2}-\d{2}$','format':'%Y-%m-%d','post':lambda x: f"{x} 00:00:00"},
                {'name':'DD/MM/YYYY','pattern':r'^\d{2}/\d{2}/\d{4}$','format':'%d/%m/%Y','post':lambda x: f"{x} 00:00:00"},
                {'name':'dayfirst','pattern':None,'dayfirst':True},
                {'name':'default', 'pattern':None},
            ]
            remaining = s.notna()
            for h in handlers:
                if not remaining.any(): break
                mask = remaining & s.str.contains(h['pattern'], regex=True, na=False) if h['pattern'] else remaining
                if not mask.any(): continue
                try:
                    to_conv = s[mask]
                    if 'preprocessor' in h: to_conv = to_conv.apply(h['preprocessor'])
                    if h.get('format') == 'ISO8601':
                        conv = pd.to_datetime(to_conv, format='ISO8601', errors='coerce')
                    elif 'format' in h:
                        conv = pd.to_datetime(to_conv, format=h['format'], errors='coerce')
                    elif 'dayfirst' in h:
                        conv = pd.to_datetime(to_conv, dayfirst=h['dayfirst'], errors='coerce')
                    else:
                        conv = pd.to_datetime(to_conv, errors='coerce')
                    if 'post' in h and conv.notna().any():
                        conv = conv.apply(h['post'])
                    ok = conv.notna()
                    if ok.any():
                        result[mask & ok] = conv[ok]
                        remaining &= ~(mask & ok)
                except Exception as e:
                    logger.error(f"Date handler '{h['name']}' error: {e}")
            try:
                dm = pd.to_datetime(result, errors='coerce').notna()
                final = result.copy()
                final[dm] = pd.to_datetime(result[dm]).dt.strftime('%Y-%m-%d %H:%M:%S')
                df[col] = final
            except Exception as e:
                logger.error(f"Final date processing error for '{col}': {e}")
                df[col] = original
        return df

    @staticmethod
    def _convert_date_column(series, logger):
        s = series.astype(str).str.strip()
        converted = pd.Series(index=series.index, dtype='object')
        bad = s.str.contains(r'1900-01-00', na=False)
        converted[bad] = '1900-01-01 00:00:00'
        valid = ~bad & s.notna() & (s != '')
        if valid.any():
            try:
                converted[valid] = pd.to_datetime(s[valid], dayfirst=True, errors='coerce')
            except Exception as e:
                logger.error(f"Date conversion error: {e}")
        fmt = converted.dt.strftime('%Y-%m-%d %H:%M:%S')
        failed = fmt.isna() & s.notna() & (s != '')
        fmt[failed] = s[failed]
        return fmt

    @staticmethod
    def clean_amount_columns(df):
        for col in df.columns:
            if not any(k in str(col).lower() for k in ['amount','debit','credit','balance','₹','value']):
                continue
            if pd.api.types.is_numeric_dtype(df[col]):
                continue
            cd = df[col].astype(str).str.strip().str.replace(r'[₹,$€£,\s]', '', regex=True)
            pm = cd.str.contains(r'\(.*\)', na=False)
            if pm.any():
                df.loc[pm, col] = '-' + cd[pm].str.replace(r'[\(\)]','',regex=True)
                cd = df[col].astype(str).str.strip()
            try:
                df[col] = pd.to_numeric(cd, errors='raise')
            except:
                df[col] = pd.to_numeric(cd.str.extract(r'([-+]?\d*\.?\d+)', expand=False), errors='coerce')
        return df

    @staticmethod
    def extract_utr_uid(df):
        nc = next((c for c in df.columns if 'narration' in str(c).lower() or 'narrative' in str(c).lower()), None)
        if nc is None: return df
        def _utr(n):
            if not isinstance(n, str): return ''
            m = re.search(r'(?:UTR|Ref No)[:/]?\s*([A-Za-z0-9]{12,22})', n, re.IGNORECASE)
            return m.group(1) if m else ''
        def _uid(n):
            if not isinstance(n, str): return ''
            m = re.search(r'(?:UID|Customer ID)[:/]?\s*([A-Za-z0-9]{8,16})', n, re.IGNORECASE)
            return m.group(1) if m else ''
        df['UTR'] = df[nc].apply(_utr)
        df['UID'] = df[nc].apply(_uid)
        return df


# ─────────────────────────────────────────────────────────────────────────────
# DatabaseManager
# ─────────────────────────────────────────────────────────────────────────────
SYSTEM_DBS = {'information_schema', 'mysql', 'performance_schema', 'sys'}

class DatabaseManager:
    def __init__(self, host, username, password, database='information_schema'):
        self.host = host
        self.username = username
        self.password = password
        self.database = database
        self.logger = LoggerManager.setup_logger()

    def _conn(self, db=None):
        return mysql.connector.connect(
            host=self.host, user=self.username, password=self.password,
            database=db or self.database,
            charset='utf8mb4', collation='utf8mb4_unicode_ci', autocommit=False
        )

    def fetch_databases(self) -> List[str]:
        try:
            c = self._conn('information_schema')
            cur = c.cursor()
            cur.execute("SHOW DATABASES")
            dbs = [r[0] for r in cur.fetchall() if r[0].lower() not in SYSTEM_DBS]
            cur.close(); c.close()
            return sorted(dbs)
        except Exception as e:
            self.logger.error(f"Error fetching databases: {e}")
            return []

    def fetch_tables(self, db_name: str) -> List[str]:
        try:
            c = self._conn(db_name)
            cur = c.cursor()
            cur.execute("SHOW TABLES")
            tables = [r[0] for r in cur.fetchall()]
            cur.close(); c.close()
            return sorted(tables)
        except Exception as e:
            self.logger.error(f"Error fetching tables for {db_name}: {e}")
            return []

    def fetch_columns(self, db_name: str, table_name: str) -> List[dict]:
        try:
            c = self._conn(db_name)
            cur = c.cursor(dictionary=True)
            cur.execute(f"""
                SELECT COLUMN_NAME, DATA_TYPE
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = '{db_name}' AND TABLE_NAME = '{table_name}'
                ORDER BY ORDINAL_POSITION
            """)
            cols = cur.fetchall()
            cur.close(); c.close()
            return cols
        except Exception as e:
            self.logger.error(f"Error fetching columns: {e}")
            return []

    def fetch_date_columns(self, db_name: str, table_name: str) -> List[str]:
        cols = self.fetch_columns(db_name, table_name)
        date_types = {'date', 'datetime', 'timestamp', 'time', 'year'}
        date_cols = [c['COLUMN_NAME'] for c in cols if c['DATA_TYPE'].lower() in date_types]
        name_cols = [c['COLUMN_NAME'] for c in cols
                     if any(k in c['COLUMN_NAME'].lower() for k in ['date','time','created','modified'])
                     and c['COLUMN_NAME'] not in date_cols]
        return date_cols + name_cols

    def fetch_all_columns(self, db_name: str, table_name: str) -> List[str]:
        cols = self.fetch_columns(db_name, table_name)
        return [c['COLUMN_NAME'] for c in cols]

    def fetch_report_data(self, db_name: str, table_name: str, date_col: str,
                          basis: str, from_date: str, to_date: str,
                          groupby_cols: List[str] = None,
                          agg_col: str = None, agg_func: str = 'COUNT') -> pd.DataFrame:
        """
        Fetch grouped/aggregated data from DB filtered by date range.
        Groups by date basis + any extra groupby_cols.
        """
        try:
            c = self._conn(db_name)
            cur = c.cursor(dictionary=True)

            # ── Date grouping expression ──────────────────────────────────────
            if basis == 'Daily':
                date_group_expr = f"DATE(`{date_col}`)"
                date_label      = f"DATE(`{date_col}`) AS `Report Date`"
            elif basis == 'Monthly':
                date_group_expr = f"DATE_FORMAT(`{date_col}`, '%Y-%m')"
                date_label      = f"DATE_FORMAT(`{date_col}`, '%Y-%m') AS `Report Month`"
            elif basis == 'Yearly':
                date_group_expr = f"YEAR(`{date_col}`)"
                date_label      = f"YEAR(`{date_col}`) AS `Report Year`"
            else:
                date_group_expr = f"DATE(`{date_col}`)"
                date_label      = f"DATE(`{date_col}`) AS `Report Date`"

            # ── Extra GROUP BY columns ────────────────────────────────────────
            extra_select_parts = []
            extra_group_parts  = []
            if groupby_cols:
                for gc in groupby_cols:
                    extra_select_parts.append(f"`{gc}`")
                    extra_group_parts.append(f"`{gc}`")

            extra_select = (", " + ", ".join(extra_select_parts)) if extra_select_parts else ""
            extra_group  = (", " + ", ".join(extra_group_parts))  if extra_group_parts  else ""

            # ── Aggregate expression ──────────────────────────────────────────
            if agg_col and agg_func and agg_func != 'COUNT' and agg_col != '(all rows)':
                agg_expr = f"{agg_func}(`{agg_col}`) AS `{agg_func}({agg_col})`"
            else:
                agg_expr = "COUNT(*) AS `Total Records`"

            query = f"""
                SELECT {date_label}{extra_select}, {agg_expr}
                FROM `{db_name}`.`{table_name}`
                WHERE `{date_col}` BETWEEN '{from_date} 00:00:00' AND '{to_date} 23:59:59'
                GROUP BY {date_group_expr}{extra_group}
                ORDER BY {date_group_expr}{extra_group}
            """
            cur.execute(query)
            rows = cur.fetchall()
            cur.close(); c.close()
            return pd.DataFrame(rows)
        except Exception as e:
            raise RuntimeError(f"Report fetch error: {e}")

    def fetch_detailed_report_data(self, db_name: str, table_name: str,
                                   date_col: str, from_date: str, to_date: str,
                                   groupby_cols: List[str] = None,
                                   limit: int = 1000) -> pd.DataFrame:
        try:
            c = self._conn(db_name)
            cur = c.cursor(dictionary=True)
            order_parts = [f"`{date_col}`"]
            if groupby_cols:
                order_parts += [f"`{g}`" for g in groupby_cols]
            order_clause = ", ".join(order_parts)
            query = f"""
                SELECT * FROM `{db_name}`.`{table_name}`
                WHERE `{date_col}` BETWEEN '{from_date} 00:00:00' AND '{to_date} 23:59:59'
                ORDER BY {order_clause}
                LIMIT {limit}
            """
            cur.execute(query)
            rows = cur.fetchall()
            cur.close(); c.close()
            return pd.DataFrame(rows)
        except Exception as e:
            raise RuntimeError(f"Detailed report fetch error: {e}")

    def create_database_if_not_exists(self, db_name: str):
        c = self._conn('information_schema')
        cur = c.cursor()
        cur.execute(f"CREATE DATABASE IF NOT EXISTS `{db_name}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
        c.commit(); cur.close(); c.close()

    def create_table_skeleton(self, db_name: str, table_name: str):
        c = self._conn(db_name)
        cur = c.cursor()
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS `{table_name}` (
                `id` INT NOT NULL AUTO_INCREMENT, PRIMARY KEY (`id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        c.commit(); cur.close(); c.close()

    def table_exists(self, db_name: str, table_name: str) -> bool:
        c = self._conn(db_name)
        cur = c.cursor()
        cur.execute(f"""
            SELECT COUNT(*) FROM information_schema.tables
            WHERE table_schema='{db_name}' AND table_name='{table_name}'
        """)
        exists = cur.fetchone()[0] == 1
        cur.close(); c.close()
        return exists

    def get_table_columns(self, db_name: str, table_name: str) -> List[str]:
        c = self._conn(db_name)
        cur = c.cursor()
        cur.execute(f"SHOW COLUMNS FROM `{db_name}`.`{table_name}`")
        cols = [r[0] for r in cur.fetchall()]
        cur.close(); c.close()
        return cols

    def get_primary_key(self, db_name: str, table_name: str) -> Optional[str]:
        c = self._conn(db_name)
        cur = c.cursor()
        cur.execute(f"""
            SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA='{db_name}' AND TABLE_NAME='{table_name}' AND COLUMN_KEY='PRI'
        """)
        r = cur.fetchone()
        cur.close(); c.close()
        return r[0] if r else None

    def export_idfc_to_excel(self, db_name: str, save_path: str) -> bool:
        c = self._conn(db_name)
        cur = c.cursor(dictionary=True)
        cur.execute("SELECT * FROM idfc")
        rows = cur.fetchall()
        cur.close(); c.close()
        if not rows: return False
        df = DataProcessor.extract_utr_uid(pd.DataFrame(rows))
        df.to_excel(save_path, index=False, engine='openpyxl')
        return True


# ─────────────────────────────────────────────────────────────────────────────
# DataUploader
# ─────────────────────────────────────────────────────────────────────────────
class DataUploader:
    def __init__(self, db_manager: DatabaseManager):
        self.db = db_manager
        self.logger = LoggerManager.setup_logger()

    def validate(self, df, db_name, table_name):
        if ConfigManager.requires_table_format_extraction(db_name, table_name):
            ec = ConfigManager.get_expected_columns(db_name, table_name)
            missing = set(ec) - set(df.columns)
            if missing: raise ValueError(f"Missing expected columns: {missing}")
        if df.empty: raise ValueError("No data to upload after processing")

    def _mysql_type(self, series):
        series = series.replace('', None)
        try:
            num = pd.to_numeric(series.dropna())
            if num.empty: return "TEXT"
            if all(num == num.astype(int)):
                mx, mn = num.max(), num.min()
                for lo, hi, typ in [(-128,127,"TINYINT"),(-32768,32767,"SMALLINT"),
                                    (-8388608,8388607,"MEDIUMINT"),(-2147483648,2147483647,"INT")]:
                    if mn >= lo and mx <= hi: return typ
                return "BIGINT"
            return "DOUBLE"
        except: pass
        try:
            parsed = pd.to_datetime(series.dropna(), errors='coerce', format='mixed')
            if parsed.notna().all(): return "DATETIME"
            raise ValueError
        except: pass
        ml = series.astype(str).str.len().max()
        if pd.isna(ml) or ml <= 255: return "VARCHAR(255)"
        if ml <= 65535: return "TEXT"
        if ml <= 16777215: return "MEDIUMTEXT"
        return "LONGTEXT"

    def _create_table(self, cursor, db_name, table_name, df):
        defs = [f"`{c}` {self._mysql_type(df[c])}" for c in df.columns]
        cursor.execute(f"""
            CREATE TABLE `{db_name}`.`{table_name}` ({', '.join(defs)})
            ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

    def _insert(self, conn, cursor, db_name, table_name, df):
        ph = ", ".join(["%s"] * len(df.columns))
        cols = ", ".join([f"`{c}`" for c in df.columns])
        q = f"INSERT INTO `{db_name}`.`{table_name}` ({cols}) VALUES ({ph})"
        batch, batch_size = [], 500
        for _, row in df.iterrows():
            batch.append(tuple(None if (pd.isna(v) or v == '') else str(v) for v in row))
            if len(batch) >= batch_size:
                try:
                    cursor.executemany(q, batch); conn.commit()
                except Exception as e:
                    self.logger.error(f"Batch error: {e}"); conn.rollback()
                    for r in batch:
                        try: cursor.execute(q, r); conn.commit()
                        except: conn.rollback()
                batch = []
        if batch:
            try: cursor.executemany(q, batch); conn.commit()
            except Exception as e:
                self.logger.error(f"Final batch error: {e}"); conn.rollback()

    def _upsert(self, conn, cursor, db_name, table_name, df, pk=None):
        if not pk:
            return self._insert(conn, cursor, db_name, table_name, df)
        cursor.execute(f"SELECT `{pk}` FROM `{db_name}`.`{table_name}`")
        existing = {r[0] for r in cursor.fetchall()}
        new_rows = df[~df[pk].isin(existing)]
        if new_rows.empty:
            self.logger.info("All records already exist – nothing new to insert")
            return
        self._insert(conn, cursor, db_name, table_name, new_rows)

    def upload(self, db_name, table_name, combined_df, log_fn):
        self.validate(combined_df, db_name, table_name)
        conn = self.db._conn(db_name)
        cursor = conn.cursor()
        if self.db.table_exists(db_name, table_name):
            log_fn(f"Table '{table_name}' exists – upserting")
            existing_cols = self.db.get_table_columns(db_name, table_name)
            pk = self.db.get_primary_key(db_name, table_name)
            for col in combined_df.columns:
                if col not in existing_cols:
                    t = self._mysql_type(combined_df[col])
                    log_fn(f"Adding column '{col}' ({t})")
                    cursor.execute(f"ALTER TABLE `{db_name}`.`{table_name}` ADD COLUMN `{col}` {t}")
                    conn.commit()
            self._upsert(conn, cursor, db_name, table_name, combined_df, pk)
        else:
            log_fn(f"Table '{table_name}' not found – creating")
            self._create_table(cursor, db_name, table_name, combined_df)
            conn.commit()
            self._insert(conn, cursor, db_name, table_name, combined_df)
        cursor.close(); conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# ReportGeneratorDialog
# ─────────────────────────────────────────────────────────────────────────────
class ReportGeneratorDialog(tk.Toplevel):
    BASES     = ["Daily", "Monthly", "Yearly"]
    FORMATS   = ["PDF Report", "Excel Spreadsheet", "Both"]
    AGG_FUNCS = ["COUNT", "SUM", "AVG", "MAX", "MIN"]

    def __init__(self, parent_gui):
        super().__init__(parent_gui.root)
        self.pg = parent_gui
        self.db_manager: Optional[DatabaseManager] = None
        self._report_df: Optional[pd.DataFrame] = None
        self._detail_df: Optional[pd.DataFrame] = None
        self._all_cols: List[str] = []
        self._groupby_checks: dict = {}          # col_name → BooleanVar

        self.title("📊 Report Generator")
        self.geometry("760x920")
        self.resizable(True, True)
        self.configure(bg="#1a1a2e")
        self.grab_set(); self.focus_force()
        self._center()
        self._build()

    def _center(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth()  - 760) // 2
        y = max(0, (self.winfo_screenheight() - 920) // 2)
        self.geometry(f"+{x}+{y}")

    # ── Build UI ──────────────────────────────────────────────────────────────
    def _build(self):
        outer = tk.Frame(self, bg="#1a1a2e")
        outer.pack(fill="both", expand=True)

        canvas  = tk.Canvas(outer, bg="#1a1a2e", highlightthickness=0)
        vscroll = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)
        vscroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self._scroll_frame    = tk.Frame(canvas, bg="#1a1a2e")
        self._scroll_frame_id = canvas.create_window((0, 0), window=self._scroll_frame, anchor="nw")

        def _on_frame_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_canvas_configure(e):
            canvas.itemconfig(self._scroll_frame_id, width=e.width)

        self._scroll_frame.bind("<Configure>", _on_frame_configure)
        canvas.bind("<Configure>", _on_canvas_configure)
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        body = self._scroll_frame

        # Header
        hdr = tk.Frame(body, bg="#16213e", pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📊  Report Generator",
                 font=("Segoe UI", 15, "bold"), bg="#16213e", fg="#e94560").pack()
        tk.Label(hdr, text="Date-wise & Group-By reports directly from your database",
                 font=("Segoe UI", 9), bg="#16213e", fg="#a8b2d8").pack()

        pad = tk.Frame(body, bg="#1a1a2e", padx=22, pady=10)
        pad.pack(fill="both", expand=True)

        # ── Section 1: DB & Table ─────────────────────────────────────────────
        self._section(pad, "① Database & Table", 0)

        r1 = tk.Frame(pad, bg="#1a1a2e"); r1.pack(fill="x", pady=(0, 5))
        self._lbl(r1, "Database:").pack(side="left", padx=(0, 6))
        self.rep_db_var = tk.StringVar()
        self.rep_db_cb  = ttk.Combobox(r1, textvariable=self.rep_db_var, state="readonly",
                                       width=24, font=("Segoe UI", 10))
        self.rep_db_cb.pack(side="left", padx=(0, 8))
        tk.Button(r1, text="🔄 Load DBs", command=self._load_dbs,
                  bg="#e94560", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=8, pady=3).pack(side="left")

        r2 = tk.Frame(pad, bg="#1a1a2e"); r2.pack(fill="x", pady=(0, 8))
        self._lbl(r2, "Table:      ").pack(side="left", padx=(0, 6))
        self.rep_tbl_var = tk.StringVar()
        self.rep_tbl_cb  = ttk.Combobox(r2, textvariable=self.rep_tbl_var, state="readonly",
                                        width=24, font=("Segoe UI", 10))
        self.rep_tbl_cb.pack(side="left", padx=(0, 8))
        tk.Button(r2, text="📋 Load Columns", command=self._load_columns,
                  bg="#0f3460", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=8, pady=3).pack(side="left")

        self.rep_db_var.trace_add('write', lambda *_: self._on_rep_db_change())

        # ── Section 2: Date Filter ────────────────────────────────────────────
        self._section(pad, "② Date Filter", 1)

        r3 = tk.Frame(pad, bg="#1a1a2e"); r3.pack(fill="x", pady=(0, 5))
        self._lbl(r3, "Date Column:").pack(side="left", padx=(0, 6))
        self.date_col_var = tk.StringVar()
        self.date_col_cb  = ttk.Combobox(r3, textvariable=self.date_col_var, state="readonly",
                                         width=28, font=("Segoe UI", 10))
        self.date_col_cb.pack(side="left")

        r4 = tk.Frame(pad, bg="#1a1a2e"); r4.pack(fill="x", pady=(0, 8))
        self._lbl(r4, "From Date: ").pack(side="left", padx=(0, 6))
        self.from_date_var = tk.StringVar(value=datetime.date.today().replace(day=1).strftime('%Y-%m-%d'))
        tk.Entry(r4, textvariable=self.from_date_var, font=("Segoe UI", 10),
                 width=13, bg="#0f3460", fg="white", insertbackground="white",
                 relief="flat").pack(side="left", padx=(0, 14))
        self._lbl(r4, "To Date:   ").pack(side="left", padx=(0, 6))
        self.to_date_var = tk.StringVar(value=datetime.date.today().strftime('%Y-%m-%d'))
        tk.Entry(r4, textvariable=self.to_date_var, font=("Segoe UI", 10),
                 width=13, bg="#0f3460", fg="white", insertbackground="white",
                 relief="flat").pack(side="left")
        self._lbl(r4, "  YYYY-MM-DD", small=True).pack(side="left", padx=(4, 0))

        # ── Section 3: Report Options ─────────────────────────────────────────
        self._section(pad, "③ Report Options", 2)

        r5 = tk.Frame(pad, bg="#1a1a2e"); r5.pack(fill="x", pady=(0, 5))
        self._lbl(r5, "Date Basis: ").pack(side="left", padx=(0, 6))
        self.basis_var = tk.StringVar(value="Daily")
        for b in self.BASES:
            tk.Radiobutton(r5, text=b, variable=self.basis_var, value=b,
                           bg="#1a1a2e", fg="#a8b2d8", selectcolor="#e94560",
                           activebackground="#1a1a2e", activeforeground="white",
                           font=("Segoe UI", 10)).pack(side="left", padx=8)

        r6 = tk.Frame(pad, bg="#1a1a2e"); r6.pack(fill="x", pady=(0, 5))
        self._lbl(r6, "Output:     ").pack(side="left", padx=(0, 6))
        self.fmt_var = tk.StringVar(value="PDF Report")
        ttk.Combobox(r6, textvariable=self.fmt_var, values=self.FORMATS,
                     state="readonly", width=22, font=("Segoe UI", 10)).pack(side="left")

        r6b = tk.Frame(pad, bg="#1a1a2e"); r6b.pack(fill="x", pady=(0, 8))
        self.include_detail = tk.BooleanVar(value=False)
        tk.Checkbutton(r6b, text="Include detailed rows (up to 1000 rows) in report",
                       variable=self.include_detail, bg="#1a1a2e", fg="#a8b2d8",
                       selectcolor="#e94560", activebackground="#1a1a2e",
                       font=("Segoe UI", 9)).pack(side="left")

        # ── Section 4: Group By ───────────────────────────────────────────────
        self._section(pad, "④ Group By  (select columns to further group the report)", 3)

        # Aggregate function + column row
        agg_row = tk.Frame(pad, bg="#1a1a2e"); agg_row.pack(fill="x", pady=(0, 6))
        self._lbl(agg_row, "Aggregate:  ").pack(side="left", padx=(0, 6))
        self.agg_func_var = tk.StringVar(value="COUNT")
        ttk.Combobox(agg_row, textvariable=self.agg_func_var, values=self.AGG_FUNCS,
                     state="readonly", width=10, font=("Segoe UI", 10)).pack(side="left", padx=(0, 10))
        self._lbl(agg_row, "of column:").pack(side="left", padx=(0, 6))
        self.agg_col_var = tk.StringVar(value="(all rows)")
        self.agg_col_cb  = ttk.Combobox(agg_row, textvariable=self.agg_col_var,
                                        state="readonly", width=22, font=("Segoe UI", 10))
        self.agg_col_cb.pack(side="left")
        self._lbl(agg_row, "  ← ignored for COUNT", small=True).pack(side="left", padx=(4, 0))

        # Group-by column checkboxes
        self._lbl(pad, "Group by these extra columns:").pack(anchor="w", pady=(2, 3))

        gb_outer = tk.Frame(pad, bg="#0d1117", relief="flat", bd=1)
        gb_outer.pack(fill="x", pady=(0, 8))

        self._gb_canvas = tk.Canvas(gb_outer, bg="#0d1117", highlightthickness=0, height=110)
        gb_scroll = ttk.Scrollbar(gb_outer, orient="vertical", command=self._gb_canvas.yview)
        self._gb_canvas.configure(yscrollcommand=gb_scroll.set)
        gb_scroll.pack(side="right", fill="y")
        self._gb_canvas.pack(side="left", fill="both", expand=True)

        self._gb_inner    = tk.Frame(self._gb_canvas, bg="#0d1117")
        self._gb_canvas_id = self._gb_canvas.create_window((0, 0), window=self._gb_inner, anchor="nw")

        def _gb_configure(e):
            self._gb_canvas.configure(scrollregion=self._gb_canvas.bbox("all"))
        self._gb_inner.bind("<Configure>", _gb_configure)

        # Quick-select buttons
        gb_btn_row = tk.Frame(pad, bg="#1a1a2e"); gb_btn_row.pack(fill="x", pady=(0, 4))
        tk.Button(gb_btn_row, text="☑ Select All", command=self._gb_select_all,
                  bg="#0f3460", fg="white", font=("Segoe UI", 8), relief="flat", padx=6, pady=2
                  ).pack(side="left", padx=(0, 4))
        tk.Button(gb_btn_row, text="☐ Clear All", command=self._gb_clear_all,
                  bg="#333", fg="#aaa", font=("Segoe UI", 8), relief="flat", padx=6, pady=2
                  ).pack(side="left")
        self.gb_selected_lbl = tk.Label(gb_btn_row, text="No columns selected",
                                        bg="#1a1a2e", fg="#e94560", font=("Segoe UI", 8))
        self.gb_selected_lbl.pack(side="left", padx=(10, 0))

        # ── Action Buttons ────────────────────────────────────────────────────
        bf = tk.Frame(pad, bg="#1a1a2e"); bf.pack(fill="x", pady=(6, 8))
        tk.Button(bf, text="🔍 Preview Report", command=self._preview_report,
                  bg="#533483", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=14, pady=7).pack(side="left", padx=(0, 8))
        tk.Button(bf, text="💾 Generate & Save", command=self._generate_report,
                  bg="#e94560", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=14, pady=7).pack(side="left", padx=(0, 8))
        tk.Button(bf, text="✖ Close", command=self.destroy,
                  bg="#333", fg="#aaa", font=("Segoe UI", 10),
                  relief="flat", padx=14, pady=7).pack(side="right")

        # ── Preview ───────────────────────────────────────────────────────────
        self._section(pad, "⑤ Preview", 0)
        self.preview_text = tk.Text(pad, height=14, bg="#0d1117", fg="#58a6ff",
                                    font=("Consolas", 9), relief="flat",
                                    padx=8, pady=6, wrap="none")
        # Horizontal scrollbar for wide tables
        px_scroll = ttk.Scrollbar(pad, orient="horizontal", command=self.preview_text.xview)
        self.preview_text.configure(xscrollcommand=px_scroll.set)
        px_scroll.pack(fill="x")
        self.preview_text.pack(fill="both", expand=True)
        self.preview_text.insert("end", "→ Load columns, configure options, then click Preview.\n")
        self.preview_text.config(state="disabled")

        # Status bar
        self.status_lbl = tk.Label(self, text="Ready", bg="#16213e", fg="#a8b2d8",
                                   font=("Segoe UI", 9), anchor="w")
        self.status_lbl.pack(fill="x", padx=10, pady=(0, 6))

    # ── Group-by checkbox helpers ─────────────────────────────────────────────
    def _rebuild_groupby_checkboxes(self, all_cols: List[str], date_col: str):
        for widget in self._gb_inner.winfo_children():
            widget.destroy()
        self._groupby_checks.clear()

        cols_to_show = [c for c in all_cols if c.lower() != (date_col.lower() if date_col else '')]

        COLS_PER_ROW = 3
        for idx, col in enumerate(cols_to_show):
            var = tk.BooleanVar(value=False)
            self._groupby_checks[col] = var
            r, c = divmod(idx, COLS_PER_ROW)
            cb = tk.Checkbutton(
                self._gb_inner, text=col, variable=var,
                command=self._update_gb_label,
                bg="#0d1117", fg="#a8b2d8", selectcolor="#e94560",
                activebackground="#0d1117", activeforeground="white",
                font=("Segoe UI", 9), anchor="w"
            )
            cb.grid(row=r, column=c, sticky="w", padx=8, pady=2)

        if not cols_to_show:
            tk.Label(self._gb_inner, text="No additional columns available.",
                     bg="#0d1117", fg="#555", font=("Segoe UI", 9)).grid(row=0, column=0, padx=8, pady=6)

        self._gb_inner.update_idletasks()
        self._gb_canvas.configure(scrollregion=self._gb_canvas.bbox("all"))
        self._update_gb_label()

    def _get_selected_groupby_cols(self) -> List[str]:
        return [col for col, var in self._groupby_checks.items() if var.get()]

    def _update_gb_label(self):
        selected = self._get_selected_groupby_cols()
        if selected:
            self.gb_selected_lbl.config(
                text=f"✔ {len(selected)} selected: {', '.join(selected[:4])}{'…' if len(selected) > 4 else ''}",
                fg="#4ecca3")
        else:
            self.gb_selected_lbl.config(text="No extra columns selected", fg="#e94560")

    def _gb_select_all(self):
        for var in self._groupby_checks.values(): var.set(True)
        self._update_gb_label()

    def _gb_clear_all(self):
        for var in self._groupby_checks.values(): var.set(False)
        self._update_gb_label()

    # ── Section / label helpers ───────────────────────────────────────────────
    def _section(self, parent, title, idx):
        colors_map = ["#e94560", "#0f3460", "#533483", "#16213e"]
        tk.Frame(parent, bg=colors_map[idx % len(colors_map)], height=2).pack(fill="x", pady=(8, 4))
        tk.Label(parent, text=title, font=("Segoe UI", 10, "bold"),
                 bg="#1a1a2e", fg="#e94560").pack(anchor="w", pady=(2, 4))

    def _lbl(self, parent, text, small=False):
        return tk.Label(parent, text=text,
                        font=("Segoe UI", 8 if small else 10),
                        bg="#1a1a2e",
                        fg="#a8b2d8" if small else "#ccd6f6")

    # ── Logic ─────────────────────────────────────────────────────────────────
    def _set_status(self, msg):
        self.status_lbl.config(text=msg)
        self.update_idletasks()

    def _get_db_manager(self):
        if self.db_manager:
            return self.db_manager
        host = self.pg.host_entry.get().strip()
        user = self.pg.username_entry.get().strip()
        pwd  = self.pg.password_entry.get()
        self.db_manager = DatabaseManager(host, user, pwd)
        return self.db_manager

    def _load_dbs(self):
        try:
            self._set_status("Connecting to server…")
            mgr = self._get_db_manager()
            dbs = mgr.fetch_databases()
            self.rep_db_cb['values'] = dbs
            self._set_status(f"✔ {len(dbs)} databases loaded.")
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)
            self._set_status(f"❌ {e}")

    def _on_rep_db_change(self):
        db = self.rep_db_var.get().strip()
        if not db: return
        try:
            mgr = self._get_db_manager()
            tables = mgr.fetch_tables(db)
            self.rep_tbl_cb['values'] = tables
            self.rep_tbl_var.set('')
            self.date_col_cb['values'] = []
            self.date_col_var.set('')
            # Clear group-by checkboxes when DB changes
            for widget in self._gb_inner.winfo_children():
                widget.destroy()
            self._groupby_checks.clear()
            self._update_gb_label()
            self._set_status(f"✔ {len(tables)} tables in '{db}'")
        except Exception as e:
            self._set_status(f"❌ {e}")

    def _load_columns(self):
        db    = self.rep_db_var.get().strip()
        table = self.rep_tbl_var.get().strip()
        if not db or not table:
            messagebox.showwarning("Select Table", "Please select a Database and Table first.", parent=self)
            return
        try:
            self._set_status("Loading columns…")
            mgr = self._get_db_manager()
            date_cols = mgr.fetch_date_columns(db, table)
            all_cols  = mgr.fetch_all_columns(db, table)
            self._all_cols = all_cols

            # Date column combobox: date cols first, then others
            combo_cols = date_cols + [c for c in all_cols if c not in date_cols]
            self.date_col_cb['values'] = combo_cols

            # Populate aggregate column dropdown
            self.agg_col_cb['values'] = ["(all rows)"] + all_cols
            self.agg_col_var.set("(all rows)")

            # Auto-select first date column
            selected_date_col = date_cols[0] if date_cols else (all_cols[0] if all_cols else "")
            if selected_date_col:
                self.date_col_var.set(selected_date_col)

            if date_cols:
                self._set_status(f"✔ {len(date_cols)} date column(s) found, {len(all_cols)} total columns.")
            else:
                self._set_status("⚠ No obvious date columns found. Choose manually.")

            # ✅ Rebuild Group By checkboxes with ALL columns (excluding selected date col)
            self._rebuild_groupby_checkboxes(all_cols, selected_date_col)

        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)
            self._set_status(f"❌ {e}")

    def _validate_inputs(self):
        db    = self.rep_db_var.get().strip()
        table = self.rep_tbl_var.get().strip()
        dcol  = self.date_col_var.get().strip()
        frm   = self.from_date_var.get().strip()
        to    = self.to_date_var.get().strip()
        if not db:    raise ValueError("Please select a Database.")
        if not table: raise ValueError("Please select a Table.")
        if not dcol:  raise ValueError("Please select a Date Column.")
        try:
            datetime.datetime.strptime(frm, '%Y-%m-%d')
            datetime.datetime.strptime(to,  '%Y-%m-%d')
        except:
            raise ValueError("Dates must be in YYYY-MM-DD format.")
        if frm > to:
            raise ValueError("'From Date' must be before or equal to 'To Date'.")
        return db, table, dcol, frm, to

    # ── Preview ───────────────────────────────────────────────────────────────
    def _preview_report(self):
        try:
            db, table, dcol, frm, to = self._validate_inputs()
        except ValueError as e:
            messagebox.showwarning("Input Error", str(e), parent=self)
            self._set_status(f"⚠ {e}")
            return

        basis        = self.basis_var.get()
        groupby_cols = self._get_selected_groupby_cols()   # ← pulled from checkboxes
        agg_func     = self.agg_func_var.get()
        agg_col      = self.agg_col_var.get()

        self._set_status("Fetching preview data…")
        try:
            mgr = self._get_db_manager()
            # Pass group-by cols to DB query
            self._report_df = mgr.fetch_report_data(
                db, table, dcol, basis, frm, to,
                groupby_cols=groupby_cols,
                agg_col=agg_col,
                agg_func=agg_func
            )
            if self.include_detail.get():
                self._detail_df = mgr.fetch_detailed_report_data(
                    db, table, dcol, frm, to,
                    groupby_cols=groupby_cols
                )
            self._show_preview(db, table, dcol, basis, frm, to, groupby_cols)
            periods = len(self._report_df)
            gb_info = f" | grouped by: {', '.join(groupby_cols)}" if groupby_cols else ""
            self._set_status(f"✔ Preview ready – {periods} row(s){gb_info}")
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)
            self._set_status(f"❌ {e}")

    def _show_preview(self, db, table, dcol, basis, frm, to, groupby_cols=None):
        self.preview_text.config(state="normal")
        self.preview_text.delete("1.0", "end")
        df = self._report_df

        gb_str = ", ".join(groupby_cols) if groupby_cols else "None"
        lines = [
            f"  Database   : {db}",
            f"  Table      : {table}",
            f"  Date Col   : {dcol}",
            f"  Basis      : {basis}",
            f"  Group By   : {gb_str}",
            f"  Period     : {frm}  →  {to}",
            f"  Result Rows: {len(df)}",
            "",
        ]

        if not df.empty:
            # ── Build a dynamic column-width table ────────────────────────────
            col_names = list(df.columns)

            # Compute max width per column (header vs data)
            col_widths = []
            for cn in col_names:
                max_data = df[cn].astype(str).str.len().max() if len(df) > 0 else 0
                col_widths.append(max(len(str(cn)), int(max_data), 8))

            # Cap individual column width at 30
            col_widths = [min(w, 30) for w in col_widths]

            sep   = "  " + "─" * (sum(col_widths) + 3 * len(col_widths))
            header_parts = [str(cn).ljust(col_widths[i]) for i, cn in enumerate(col_names)]
            header_line  = "  " + "  │  ".join(header_parts)

            lines.append(sep)
            lines.append(header_line)
            lines.append(sep)

            for _, row in df.iterrows():
                row_parts = [str(row[cn])[:col_widths[i]].ljust(col_widths[i])
                             for i, cn in enumerate(col_names)]
                lines.append("  " + "  │  ".join(row_parts))

            lines.append(sep)

            # Totals row: sum numeric columns (last column typically)
            total_parts = []
            for i, cn in enumerate(col_names):
                numeric_total = pd.to_numeric(df[cn], errors='coerce')
                if numeric_total.notna().any():
                    total_val = str(int(numeric_total.sum()))
                    total_parts.append(("TOTAL" if i == 0 else total_val).ljust(col_widths[i]))
                else:
                    total_parts.append(("TOTAL" if i == 0 else "").ljust(col_widths[i]))
            lines.append("  " + "  │  ".join(total_parts))
            lines.append(sep)
        else:
            lines.append("  ⚠  No data found for the selected period / group-by combination.")

        self.preview_text.insert("end", "\n".join(lines))
        self.preview_text.config(state="disabled")

    # ── Generate & Save ───────────────────────────────────────────────────────
    def _generate_report(self):
        try:
            db, table, dcol, frm, to = self._validate_inputs()
        except ValueError as e:
            messagebox.showwarning("Input Error", str(e), parent=self)
            return

        basis        = self.basis_var.get()
        groupby_cols = self._get_selected_groupby_cols()
        agg_func     = self.agg_func_var.get()
        agg_col      = self.agg_col_var.get()

        if self._report_df is None or self._report_df.empty:
            try:
                mgr = self._get_db_manager()
                self._report_df = mgr.fetch_report_data(
                    db, table, dcol, basis, frm, to,
                    groupby_cols=groupby_cols, agg_col=agg_col, agg_func=agg_func
                )
                if self.include_detail.get():
                    self._detail_df = mgr.fetch_detailed_report_data(
                        db, table, dcol, frm, to, groupby_cols=groupby_cols
                    )
            except Exception as e:
                messagebox.showerror("Fetch Error", str(e), parent=self)
                return

        if self._report_df.empty:
            messagebox.showinfo("No Data", "No records found for the selected date range.", parent=self)
            return

        fmt      = self.fmt_var.get()
        save_dir = filedialog.askdirectory(title="Select folder to save report", parent=self)
        if not save_dir: return

        ts        = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        base_name = f"Report_{db}_{table}_{basis}_{frm}_to_{to}_{ts}"
        saved     = []

        try:
            if fmt in ["PDF Report", "Both"]:
                pdf_path = os.path.join(save_dir, base_name + ".pdf")
                self._save_pdf(pdf_path, db, table, dcol, basis, frm, to, groupby_cols)
                saved.append(f"PDF: {pdf_path}")

            if fmt in ["Excel Spreadsheet", "Both"]:
                xl_path = os.path.join(save_dir, base_name + ".xlsx")
                self._save_excel(xl_path, db, table, dcol, basis, frm, to, groupby_cols)
                saved.append(f"Excel: {xl_path}")

            msg = "Report(s) saved:\n\n" + "\n".join(saved)
            messagebox.showinfo("✅ Report Saved", msg, parent=self)
            self._set_status(f"✔ Saved: {', '.join(saved)}")
            self.pg.log_status(f"📊 Report generated: {', '.join(saved)}")
        except Exception as e:
            messagebox.showerror("Save Error", str(e), parent=self)
            self._set_status(f"❌ {e}")

    # ── PDF Generation ────────────────────────────────────────────────────────
    def _save_pdf(self, path, db, table, dcol, basis, frm, to, groupby_cols=None):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors as rl_colors

        df = self._report_df
        gb_str = ", ".join(groupby_cols) if groupby_cols else "None"

        # Landscape for group-by or detail tables
        use_landscape = bool(groupby_cols) or (
            self.include_detail.get() and self._detail_df is not None and not self._detail_df.empty
        ) or len(df.columns) > 4

        pagesize = landscape(A4) if use_landscape else A4

        doc = SimpleDocTemplate(path, pagesize=pagesize,
                                rightMargin=1.5*cm, leftMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title', parent=styles['Title'],
                                     fontSize=16, textColor=rl_colors.HexColor('#e94560'),
                                     spaceAfter=6)
        sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                   fontSize=9, textColor=rl_colors.HexColor('#555555'),
                                   spaceAfter=4)
        heading_style = ParagraphStyle('heading', parent=styles['Heading2'],
                                       fontSize=11, textColor=rl_colors.HexColor('#1a1a2e'),
                                       spaceAfter=4)

        story = []

        # Title
        story.append(Paragraph(f"{basis} Report — {db}.{table}", title_style))
        story.append(Paragraph(
            f"Date Column: <b>{dcol}</b>  |  Period: <b>{frm}</b> to <b>{to}</b>  |  "
            f"Group By: <b>{gb_str}</b>  |  "
            f"Generated: <b>{datetime.datetime.now().strftime('%d-%m-%Y %H:%M')}</b>", sub_style))
        story.append(HRFlowable(width="100%", thickness=2, color=rl_colors.HexColor('#e94560')))
        story.append(Spacer(1, 10))

        # Summary Table
        gb_title = f" (Grouped by: {gb_str})" if groupby_cols else ""
        story.append(Paragraph(f"{basis} Summary{gb_title}", heading_style))

        # Build table data from all columns in df
        col_names = list(df.columns)
        data = [col_names]
        for _, row in df.iterrows():
            data.append([str(row[c]) if pd.notna(row[c]) else '' for c in col_names])

        # Totals row
        total_row = []
        for i, cn in enumerate(col_names):
            numeric_vals = pd.to_numeric(df[cn], errors='coerce')
            if numeric_vals.notna().any():
                total_row.append("TOTAL" if i == 0 else str(int(numeric_vals.sum())))
            else:
                total_row.append("TOTAL" if i == 0 else "")
        data.append(total_row)

        pw = pagesize[0] - 3*cm
        num_cols = len(col_names)
        col_w = pw / num_cols

        tbl = Table(data, colWidths=[col_w]*num_cols, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0),  (-1, 0),  rl_colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR',     (0, 0),  (-1, 0),  rl_colors.white),
            ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0),  (-1, 0),  9),
            ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [rl_colors.HexColor('#f5f7fa'), rl_colors.white]),
            ('BACKGROUND',    (0, -1), (-1, -1), rl_colors.HexColor('#e94560')),
            ('TEXTCOLOR',     (0, -1), (-1, -1), rl_colors.white),
            ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 1),  (-1, -1), 8),
            ('ALIGN',         (0, 0),  (-1, -1), 'LEFT'),
            ('ALIGN',         (1, 0),  (-1, -1), 'RIGHT'),
            ('GRID',          (0, 0),  (-1, -1), 0.5, rl_colors.HexColor('#dddddd')),
            ('TOPPADDING',    (0, 0),  (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0),  (-1, -1), 4),
            ('LEFTPADDING',   (0, 0),  (-1, -1), 6),
        ]))
        story.append(tbl)

        # Detail table
        if self.include_detail.get() and self._detail_df is not None and not self._detail_df.empty:
            story.append(Spacer(1, 16))
            story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor('#cccccc')))
            story.append(Spacer(1, 6))
            detail_df = self._detail_df
            story.append(Paragraph("Detailed Records (up to 1000 rows)", heading_style))

            dcols     = list(detail_df.columns)
            max_cols  = min(len(dcols), 10)
            dcols     = dcols[:max_cols]
            d_data    = [dcols]
            for _, row in detail_df.iterrows():
                d_data.append([str(row[c])[:28] if pd.notna(row[c]) else '' for c in dcols])

            col_w2 = pw / max_cols
            dtbl   = Table(d_data, colWidths=[col_w2]*max_cols, repeatRows=1)
            dtbl.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0),  (-1, 0),  rl_colors.HexColor('#0f3460')),
                ('TEXTCOLOR',     (0, 0),  (-1, 0),  rl_colors.white),
                ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0),  (-1, -1), 7),
                ('ROWBACKGROUNDS',(0, 1),  (-1, -1), [rl_colors.HexColor('#f5f7fa'), rl_colors.white]),
                ('GRID',          (0, 0),  (-1, -1), 0.3, rl_colors.HexColor('#cccccc')),
                ('ALIGN',         (0, 0),  (-1, -1), 'LEFT'),
                ('TOPPADDING',    (0, 0),  (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0),  (-1, -1), 3),
                ('LEFTPADDING',   (0, 0),  (-1, -1), 4),
            ]))
            story.append(dtbl)

        # Footer
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor('#dddddd')))
        story.append(Paragraph(
            f"MySQL Excel Data Uploader  •  Auto-generated on "
            f"{datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')}",
            ParagraphStyle('footer', parent=styles['Normal'],
                           fontSize=7, textColor=rl_colors.HexColor('#999999'))
        ))
        doc.build(story)

    # ── Excel Generation ──────────────────────────────────────────────────────
    def _save_excel(self, path, db, table, dcol, basis, frm, to, groupby_cols=None):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        gb_str     = ", ".join(groupby_cols) if groupby_cols else "None"
        wb         = openpyxl.Workbook()
        dark_fill  = PatternFill("solid", fgColor="1A1A2E")
        red_fill   = PatternFill("solid", fgColor="E94560")
        blue_fill  = PatternFill("solid", fgColor="0F3460")
        alt_fill   = PatternFill("solid", fgColor="F5F7FA")
        thin = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )

        # ── Summary Sheet ──────────────────────────────────────────────────────
        ws         = wb.active
        ws.title   = f"{basis} Summary"
        df         = self._report_df
        col_names  = list(df.columns)
        num_cols   = len(col_names)
        merge_end  = get_column_letter(max(num_cols, 2))

        ws.merge_cells(f'A1:{merge_end}1')
        ws['A1'] = f"Report: {db}.{table} — {basis} Basis | Group By: {gb_str}"
        ws['A1'].font      = Font(bold=True, size=13, color="FFFFFF")
        ws['A1'].fill      = dark_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22

        ws.merge_cells(f'A2:{merge_end}2')
        ws['A2'] = (f"Date Column: {dcol}  |  Period: {frm} to {to}  |  "
                    f"Generated: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M')}")
        ws['A2'].font      = Font(size=9, color="555555")
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 16

        # Column headers (row 4)
        for ci, h in enumerate(col_names, start=1):
            cell            = ws.cell(row=4, column=ci, value=h)
            cell.font       = Font(bold=True, color="FFFFFF")
            cell.fill       = blue_fill
            cell.alignment  = Alignment(horizontal='center')
            cell.border     = thin
        ws.row_dimensions[4].height = 18

        # Data rows
        for ri, (_, row) in enumerate(df.iterrows(), start=5):
            fill = alt_fill if ri % 2 == 0 else PatternFill()
            for ci, cn in enumerate(col_names, start=1):
                val  = row[cn]
                cell = ws.cell(row=ri, column=ci, value=(str(val) if pd.notna(val) else ''))
                cell.fill      = fill
                cell.border    = thin
                cell.alignment = Alignment(horizontal='right' if ci > 1 else 'left')

        # Totals row
        tr = 5 + len(df)
        for ci, cn in enumerate(col_names, start=1):
            numeric_vals = pd.to_numeric(df[cn], errors='coerce')
            if numeric_vals.notna().any():
                val_str = "TOTAL" if ci == 1 else int(numeric_vals.sum())
            else:
                val_str = "TOTAL" if ci == 1 else ""
            cell            = ws.cell(row=tr, column=ci, value=val_str)
            cell.font       = Font(bold=True, color="FFFFFF")
            cell.fill       = red_fill
            cell.border     = thin
            cell.alignment  = Alignment(horizontal='right' if ci > 1 else 'left')

        # Auto-width
        for ci in range(1, num_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 22

        # ── Detail Sheet ───────────────────────────────────────────────────────
        if self.include_detail.get() and self._detail_df is not None and not self._detail_df.empty:
            ds         = wb.create_sheet(title="Detailed Records")
            detail_df  = self._detail_df
            dcols      = list(detail_df.columns)

            for ci, h in enumerate(dcols, start=1):
                cell            = ds.cell(row=1, column=ci, value=h)
                cell.font       = Font(bold=True, color="FFFFFF")
                cell.fill       = dark_fill
                cell.border     = thin
                cell.alignment  = Alignment(horizontal='center')

            for ri, (_, row) in enumerate(detail_df.iterrows(), start=2):
                fill = alt_fill if ri % 2 == 0 else PatternFill()
                for ci, col in enumerate(dcols, start=1):
                    val  = row[col]
                    cell = ds.cell(row=ri, column=ci, value=(str(val) if pd.notna(val) else ''))
                    cell.fill   = fill
                    cell.border = thin

            for ci in range(1, len(dcols)+1):
                ds.column_dimensions[get_column_letter(ci)].width = 18

        wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# CreateDBTableDialog
# ─────────────────────────────────────────────────────────────────────────────
class CreateDBTableDialog(tk.Toplevel):
    def __init__(self, parent_gui):
        super().__init__(parent_gui.root)
        self.pg = parent_gui
        self.title("➕ Create New Database / Table")
        self.geometry("480x400")
        self.resizable(False, False)
        self.configure(bg="#f5f7fa")
        self.grab_set(); self.focus_force()
        self._center(); self._build()

    def _center(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth()-480)//2
        y = (self.winfo_screenheight()-400)//2
        self.geometry(f"+{x}+{y}")

    def _build(self):
        tk.Label(self, text="Create New Database / Table",
                 font=("Segoe UI",13,"bold"), bg="#f5f7fa", fg="#2c3e50").pack(pady=(16,2))
        tk.Label(self, text="Leave 'New Database' blank to use an existing one.",
                 font=("Segoe UI",9), bg="#f5f7fa", fg="#7f8c8d").pack()
        ttk.Separator(self).pack(fill="x", padx=18, pady=10)

        f = tk.Frame(self, bg="#f5f7fa"); f.pack(fill="both", padx=18)
        p = {"padx":10,"pady":6}

        tk.Label(f, text="New Database Name:", bg="#f5f7fa", font=("Segoe UI",10)).grid(row=0,column=0,sticky="w",**p)
        self.new_db = tk.StringVar()
        tk.Entry(f, textvariable=self.new_db, font=("Segoe UI",10), width=28).grid(row=0,column=1,sticky="ew",**p)

        tk.Label(f, text="— OR pick existing —", bg="#f5f7fa", fg="#95a5a6",
                 font=("Segoe UI",9)).grid(row=1,column=0,columnspan=3,pady=(0,2))

        tk.Label(f, text="Existing Database:", bg="#f5f7fa", font=("Segoe UI",10)).grid(row=2,column=0,sticky="w",**p)
        self.exist_db = tk.StringVar()
        self.exist_cb = ttk.Combobox(f, textvariable=self.exist_db, font=("Segoe UI",10), width=25, state="readonly")
        self.exist_cb.grid(row=2,column=1,sticky="ew",**p)
        tk.Button(f, text="🔄 Fetch", font=("Segoe UI",9), command=self._fetch,
                  relief="flat", bg="#3498db", fg="white", padx=6).grid(row=2,column=2,padx=(0,6))

        ttk.Separator(f,orient="horizontal").grid(row=3,column=0,columnspan=3,sticky="ew",pady=8)

        tk.Label(f, text="New Table Name:*", bg="#f5f7fa", font=("Segoe UI",10)).grid(row=4,column=0,sticky="w",**p)
        self.new_tbl = tk.StringVar()
        tk.Entry(f, textvariable=self.new_tbl, font=("Segoe UI",10), width=28).grid(row=4,column=1,sticky="ew",**p)
        f.columnconfigure(1, weight=1)

        self.status = tk.Label(self, text="", bg="#f5f7fa", font=("Segoe UI",9), fg="#e74c3c", wraplength=440)
        self.status.pack(pady=4)

        bf = tk.Frame(self, bg="#f5f7fa"); bf.pack(pady=8)
        tk.Button(bf, text="✅ Create", font=("Segoe UI",10,"bold"),
                  bg="#27ae60", fg="white", padx=14, pady=6, relief="flat",
                  command=self._create).pack(side="left", padx=6)
        tk.Button(bf, text="Cancel", font=("Segoe UI",10),
                  bg="#e74c3c", fg="white", padx=14, pady=6, relief="flat",
                  command=self.destroy).pack(side="left", padx=6)

    def _fetch(self):
        try:
            mgr = DatabaseManager(self.pg.host_entry.get().strip(),
                                  self.pg.username_entry.get().strip(),
                                  self.pg.password_entry.get())
            dbs = mgr.fetch_databases()
            self.exist_cb['values'] = dbs
            self.status.config(text=f"✔ {len(dbs)} databases loaded.", fg="#27ae60")
        except Exception as e:
            self.status.config(text=f"Connection error: {e}", fg="#e74c3c")

    def _create(self):
        new_db  = self.new_db.get().strip()
        exist   = self.exist_db.get().strip()
        new_tbl = self.new_tbl.get().strip()
        target  = new_db or exist
        if not target:
            self.status.config(text="Provide a new or existing database name."); return
        if not new_tbl:
            self.status.config(text="Table name is required."); return
        if not re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', target):
            self.status.config(text="Database name: letters/digits/_ only."); return
        if not re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', new_tbl):
            self.status.config(text="Table name: letters/digits/_ only."); return
        if not messagebox.askyesno("Confirm", f"Create database '{target}' (if new) and table '{new_tbl}'?", parent=self):
            return
        try:
            mgr = DatabaseManager(self.pg.host_entry.get().strip(),
                                  self.pg.username_entry.get().strip(),
                                  self.pg.password_entry.get())
            mgr.create_database_if_not_exists(target)
            mgr.create_table_skeleton(target, new_tbl)
            self.pg._refresh_db_dropdown()
            self.pg.db_var.set(target)
            self.pg._load_tables(target)
            self.pg.table_var.set(new_tbl)
            self.pg.update_instructions()
            self.pg.log_status(f"✅ Created database '{target}' and table '{new_tbl}'")
            messagebox.showinfo("Success", f"Created:\n  Database: {target}\n  Table: {new_tbl}", parent=self)
            self.destroy()
        except Exception as e:
            self.status.config(text=f"Error: {e}", fg="#e74c3c")


# ─────────────────────────────────────────────────────────────────────────────
# MySQLExcelUploaderGUI
# ─────────────────────────────────────────────────────────────────────────────
class MySQLExcelUploaderGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("MySQL Excel Data Uploader")
        self.root.geometry("980x780")
        self.logger = LoggerManager.setup_logger()
        self.logger.info("Application starting...")
        self.db_manager: Optional[DatabaseManager] = None
        self.setup_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.root.update()
        self.check_password()

    # ── UI ────────────────────────────────────────────────────────────────────
    def setup_ui(self):
        style = ttk.Style()
        style.theme_use('default')
        style.configure("TFrame",      background="#f5f7fa")
        style.configure("TLabelFrame", background="#ffffff", borderwidth=2, relief="solid", padding=10)
        style.configure("TLabelFrame.Label", background="#ffffff", font=("Segoe UI",10,"bold"))
        style.configure("TLabel",  background="#ffffff", font=("Segoe UI",10))
        style.configure("TEntry",  font=("Segoe UI",10), padding=8, fieldbackground="#f8f9fa")
        for name, bg, abg in [("Primary","#4CAF50","#45a049"),("Secondary","#3498db","#2980b9"),
                               ("Create","#8e44ad","#7d3c98"),("Fetch","#e67e22","#d35400"),
                               ("Report","#e94560","#c73652")]:
            style.configure(f"{name}.TButton", background=bg, foreground="white",
                            font=("Segoe UI",10,"bold"), padding=8, borderwidth=0)
            style.map(f"{name}.TButton", background=[("active",abg)])

        main = ttk.Frame(self.root, padding="20")
        main.pack(fill='both', expand=True)

        # ── Connection ────────────────────────────────────────────────────────
        cf = ttk.LabelFrame(main, text="Database Connection", padding="15")
        cf.pack(fill='x', pady=(0,10))

        for row, (lbl, default, show) in enumerate([
            ("Host:", "192.168.1.11", "-"),
            ("Username:", "SRA", "%"),
            ("Password:", "123", "*"),
        ]):
            ttk.Label(cf, text=lbl).grid(row=row, column=0, sticky='w', padx=5, pady=5)
            e = ttk.Entry(cf, show=show); e.insert(0, default)
            e.grid(row=row, column=1, columnspan=3, sticky='ew', padx=5, pady=5)
            setattr(self, ['host_entry','username_entry','password_entry'][row], e)

        ttk.Label(cf, text="Database:").grid(row=3, column=0, sticky='w', padx=5, pady=5)
        self.db_var = tk.StringVar()
        self.db_cb  = ttk.Combobox(cf, textvariable=self.db_var, state="readonly", width=30)
        self.db_cb.grid(row=3, column=1, sticky='ew', padx=5, pady=5)
        ttk.Button(cf, text="🔄 Load DBs", style="Fetch.TButton",
                   command=self._refresh_db_dropdown
                   ).grid(row=3, column=2, padx=(4,2), pady=5)
        ttk.Button(cf, text="➕ Create New DB / Table", style="Create.TButton",
                   command=lambda: CreateDBTableDialog(self)
                   ).grid(row=3, column=3, padx=(2,5), pady=5)

        ttk.Label(cf, text="Table:").grid(row=4, column=0, sticky='w', padx=5, pady=5)
        self.table_var = tk.StringVar()
        self.table_cb  = ttk.Combobox(cf, textvariable=self.table_var, state="readonly", width=30)
        self.table_cb.grid(row=4, column=1, sticky='ew', padx=5, pady=5)

        self.db_var.trace_add('write',    lambda *_: self._on_db_selected())
        self.table_var.trace_add('write', lambda *_: self.update_instructions())

        self.instr_var = tk.StringVar(value="Select a database, then a table.")
        ttk.Label(cf, text="Instructions:").grid(row=5, column=0, sticky='nw', padx=5, pady=5)
        ttk.Label(cf, textvariable=self.instr_var, wraplength=560, foreground="#2c3e50",
                  font=("Segoe UI",9)).grid(row=5, column=1, columnspan=3, sticky='w', padx=5, pady=5)
        cf.columnconfigure(1, weight=1)

        # ── File Upload ───────────────────────────────────────────────────────
        uf = ttk.LabelFrame(main, text="File Upload", padding="15")
        uf.pack(fill='x', pady=(0,10))

        ttk.Label(uf, text="Select File:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.file_var = tk.StringVar()
        ttk.Entry(uf, textvariable=self.file_var, state='readonly'
                  ).grid(row=0, column=1, sticky='ew', padx=5, pady=5)
        ttk.Button(uf, text="Browse", command=self.browse_file
                   ).grid(row=0, column=2, padx=5, pady=5)

        bf = ttk.Frame(uf); bf.grid(row=1, column=0, columnspan=3, sticky='ew', pady=(5,0))
        ttk.Button(bf, text="⬆ Upload Excel File",
                   command=self.initiate_upload, style="Primary.TButton"
                   ).pack(side='left', expand=True, fill='x', padx=2)
        ttk.Button(bf, text="📄 Generate Upload PDF",
                   command=self.generate_pdf_button_handler, style="Secondary.TButton"
                   ).pack(side='left', expand=True, fill='x', padx=2)
        ttk.Button(bf, text="📊 Reports Generated",
                   command=self._open_report_generator, style="Report.TButton"
                   ).pack(side='left', expand=True, fill='x', padx=2)

        uf.columnconfigure(1, weight=1)

        # ── Status ────────────────────────────────────────────────────────────
        sf = ttk.LabelFrame(main, text="Upload Status", padding="15")
        sf.pack(fill='both', expand=True)
        tf = ttk.Frame(sf); tf.pack(fill='both', expand=True)
        self.status_text = tk.Text(tf, wrap='word', font=('Consolas',10),
                                   padx=10, pady=10, bg="#f8f9fa", relief="flat",
                                   borderwidth=1, highlightthickness=1,
                                   highlightbackground="#dfe6e9")
        sb = ttk.Scrollbar(tf, orient='vertical', command=self.status_text.yview)
        self.status_text.configure(yscrollcommand=sb.set)
        self.status_text.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')
        ttk.Button(sf, text="Clear Log", command=self.clear_log).pack(pady=(10,5))

    def _open_report_generator(self):
        ReportGeneratorDialog(self)

    def _refresh_db_dropdown(self):
        host = self.host_entry.get().strip()
        user = self.username_entry.get().strip()
        pwd  = self.password_entry.get()
        if not host or not user:
            messagebox.showerror("Error", "Please enter Host and Username first.")
            return
        try:
            self.db_manager = DatabaseManager(host, user, pwd)
            dbs = self.db_manager.fetch_databases()
            if dbs:
                self.db_cb['values'] = dbs
                self.log_status(f"✔ Loaded {len(dbs)} databases.")
            else:
                self.log_status("⚠ No user databases found.")
        except Exception as e:
            messagebox.showerror("Connection Error", str(e))
            self.log_status(f"❌ Connection failed: {e}")

    def _on_db_selected(self):
        db = self.db_var.get().strip()
        if not db: return
        self._load_tables(db)

    def _load_tables(self, db_name: str):
        if not self.db_manager:
            host = self.host_entry.get().strip()
            user = self.username_entry.get().strip()
            pwd  = self.password_entry.get()
            self.db_manager = DatabaseManager(host, user, pwd)
        try:
            tables = self.db_manager.fetch_tables(db_name)
            self.table_cb['values'] = tables
            self.table_var.set('')
            if tables:
                self.log_status(f"✔ Loaded {len(tables)} tables from '{db_name}'.")
            else:
                self.log_status(f"⚠ No tables found in '{db_name}'.")
        except Exception as e:
            self.log_status(f"❌ Failed to load tables: {e}")

    def update_instructions(self):
        db    = self.db_var.get().strip()
        table = self.table_var.get().strip()
        if db and table:
            self.instr_var.set(ConfigManager.get_table_instructions(db, table))
        else:
            self.instr_var.set("Select a database, then a table.")

    def on_close(self):
        self.root.destroy()

    def _get_db_table(self):
        db    = self.db_var.get().strip()
        table = self.table_var.get().strip()
        if not db or not table:
            raise ValueError("Please select both a Database and a Table.")
        return db, table

    def browse_file(self):
        try:
            db, table = self._get_db_table()
        except ValueError as e:
            messagebox.showerror("Error", str(e)); return
        initial_dir = self._file_location(db, table)
        if not os.path.exists(initial_dir):
            initial_dir = os.getcwd()
        paths = filedialog.askopenfilenames(
            title=f"Select Files for {db}.{table}",
            initialdir=initial_dir,
            filetypes=[("Excel/CSV","*.xlsx *.xls *.csv"),("All Files","*.*")]
        )
        if paths:
            self.file_var.set("; ".join(paths))
            self.log_status(f"Selected {len(paths)} file(s)")

    def initiate_upload(self):
        try:
            db, table = self._get_db_table()
        except ValueError as e:
            messagebox.showerror("Error", str(e)); return

        file_paths = self.file_var.get()
        if not file_paths:
            messagebox.showerror("Error", "Please select files to upload."); return

        file_list = file_paths.split("; ") if "; " in file_paths else [file_paths]
        for fp in file_list:
            if not os.path.exists(fp):
                messagebox.showerror("Error", f"File not found:\n{fp}"); return
            if not fp.lower().endswith(('.xlsx','.xls','.csv')):
                messagebox.showerror("Error", f"Invalid file type:\n{fp}"); return

        host = self.host_entry.get().strip()
        user = self.username_entry.get().strip()
        pwd  = self.password_entry.get()
        if not host or not user:
            messagebox.showerror("Error", "Please fill in Host and Username."); return

        self.log_status("=" * 50)
        self.log_status(f"Starting upload for {len(file_list)} file(s)…")
        self.log_status(f"Target: {db}.{table}")
        self.log_status("=" * 50)

        threading.Thread(
            target=self._upload_thread,
            args=(host, user, pwd, db, table, file_list),
            daemon=True
        ).start()

    def _upload_thread(self, host, user, pwd, db, table, file_list):
        try:
            mgr = DatabaseManager(host, user, pwd)
            mgr.create_database_if_not_exists(db)
            uploader = DataUploader(mgr)
            combined = None

            for i, fp in enumerate(file_list, 1):
                self.log_status(f"Processing file {i}/{len(file_list)}: {os.path.basename(fp)}")
                df = (pd.read_csv(fp, dtype=str, keep_default_na=False)
                      if fp.lower().endswith('.csv')
                      else pd.read_excel(fp, dtype=str, keep_default_na=False, engine='openpyxl'))
                processed = DataProcessor.process_data(df, db, table, self.logger)
                combined  = processed if combined is None else pd.concat([combined, processed], ignore_index=True)
                self.log_status(f"  → {len(processed)} rows from this file")

            if combined is not None:
                self.log_status(f"Uploading {len(combined)} total rows…")
                uploader.upload(db, table, combined, self.log_status)
                self.log_status("=" * 50)
                self.log_status(f"✅ {len(file_list)} FILE(S) UPLOADED SUCCESSFULLY!")
                self.log_status(f"Target Database: {db}")
                self.log_status(f"Target Table: {table}")
                self.log_status(f"Total rows uploaded: {len(combined)}")
                self.log_status("=" * 50)
                if db.lower() == 'attroidfc' and table.lower() == 'idfc':
                    self.root.after(0, self._offer_idfc_export, db)
            else:
                self.log_status("No valid data found in any files.")
        except Exception as e:
            self.log_status("=" * 50)
            self.log_status(f"❌ UPLOAD FAILED: {e}")
            self.log_status("=" * 50)
            self.root.after(0, lambda: messagebox.showerror("Upload Failed", str(e)))

    def _offer_idfc_export(self, db):
        if messagebox.askyesno("Export IDFC", "Export IDFC data to Excel?"):
            path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel","*.xlsx")])
            if path:
                try:
                    mgr = DatabaseManager(self.host_entry.get().strip(),
                                          self.username_entry.get().strip(),
                                          self.password_entry.get())
                    if mgr.export_idfc_to_excel(db, path):
                        messagebox.showinfo("Exported", f"Saved to:\n{path}")
                except Exception as e:
                    messagebox.showerror("Export Failed", str(e))

    def generate_pdf_button_handler(self):
        try:
            db, table = self._get_db_table()
        except ValueError as e:
            messagebox.showerror("Error", str(e)); return
        log_lines = self.status_text.get("1.0", tk.END).strip().splitlines()
        self.generate_pdf_report(db, table, log_lines)

    def generate_pdf_report(self, database, table, log_lines):
        try:
            today    = datetime.datetime.today().strftime('%d-%m-%Y')
            save_dir = r"C:\Users\Lenovo\OneDrive\Desktop\Rajan\Database Data"
            os.makedirs(save_dir, exist_ok=True)
            save_path = os.path.join(save_dir, f"Upload_Report_{today}.pdf")
            keywords  = ["FILES UPLOADED SUCCESSFULLY","FILE(S) UPLOADED SUCCESSFULLY",
                         "Target Database:","Target Table:","Total rows uploaded:",
                         "UPLOAD FAILED:","No valid data found"]
            filtered  = []
            for line in log_lines:
                cl = line.strip()
                if cl.startswith('['):
                    be = cl.find(']')
                    if be != -1: cl = cl[be+1:].strip()
                if any(k in cl for k in keywords):
                    filtered.append(cl)
            if not filtered:
                messagebox.showinfo("No Data","No relevant information to generate a PDF report.")
                return
            c = rl_canvas.Canvas(save_path, pagesize=A4)
            w, h = A4; lm, lh, ss, bm = 60, 25, 35, 80; y = h - 80
            def pb(needed):
                nonlocal y
                if y - needed < bm: c.showPage(); y = h - 80
            c.setFont("Helvetica-Bold",18); c.drawString(lm,y,"UPLOAD SUMMARY REPORT"); y -= 15
            c.line(lm,y,w-60,y); y -= ss
            c.setFont("Helvetica",12); c.drawString(lm,y,f"Report Generated: {today}"); y -= ss
            for line in filtered:
                if "UPLOADED SUCCESSFULLY" in line:
                    pb(ss); c.setFont("Helvetica-Bold",14); c.setFillColorRGB(0,0.7,0)
                    c.drawString(lm,y,"+ "+line); c.setFillColorRGB(0,0,0); y -= ss
                elif any(k in line for k in ["Target Database:","Target Table:","Total rows uploaded:"]):
                    pb(lh); c.setFont("Helvetica",11); c.drawString(lm+20,y,f"- {line}"); y -= lh
                elif any(k in line for k in ["UPLOAD FAILED:","No valid data found"]):
                    pb(lh*2); c.setFont("Helvetica-Bold",12); c.setFillColorRGB(0.8,0,0)
                    c.drawString(lm,y,"UPLOAD FAILED"); y -= lh
                    c.setFont("Helvetica",11); c.drawString(lm+20,y,line)
                    c.setFillColorRGB(0,0,0); y -= lh
            c.setFont("Helvetica",10); c.setFillColorRGB(0.5,0.5,0.5)
            c.drawString(lm, 60, f"Generated on {datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
            c.save()
            self.log_status(f"PDF saved: {save_path}")
            messagebox.showinfo("PDF Generated", f"Saved:\n{save_path}")
        except Exception as e:
            self.log_status(f"PDF failed: {e}")

    @staticmethod
    def _file_location(db, table):
        locations = {
            'vibepay_Portal.vibepay_payin_report':      r"\\server\F DRIVE DESKTOP\Parties Record\Vibepay\FY 2025-26\01 Statements and Records\01 Vibepay Payin Data\Payin Report",
            'vibepay_Portal.vibepay_payin_wallet':      r"\\server\F DRIVE DESKTOP\Parties Record\Vibepay\FY 2025-26\01 Statements and Records\01 Vibepay Payin Data\Payin Wallet",
            'vibepay_Portal.vibepay_payout_report':     r"\\server\F DRIVE DESKTOP\Parties Record\Vibepay\FY 2025-26\01 Statements and Records\02 Vibepay Payout Data\Payout Report (Daily)",
            'vibepay_Portal.vibepay_payout_wallet':     r"\\server\F DRIVE DESKTOP\Parties Record\Vibepay\FY 2025-26\01 Statements and Records\02 Vibepay Payout Data\Payout Wallet",
            'vibepay_Portal.vibepay_add_fund_report':   r"\\server\F DRIVE DESKTOP\Parties Record\Vibepay\FY 2025-26\01 Statements and Records\05 Vibepay Add Fund Report",
            'vibepay_Portal.vibepayin_add_fund_report': r"\\server\F DRIVE DESKTOP\Parties Record\Vibepay\FY 2025-26\01 Statements and Records\01 Vibepay Payin Data\Add Fund Report",
            'sabpaisa.payin':             r"\\server\F DRIVE DESKTOP\Parties Record\Vibepay\FY 2025-26\01 Statements and Records\03 Subpaisa Payin Report\Transaction Report",
            'sabpaisa.settel_hai':        r"\\server\F DRIVE DESKTOP\Parties Record\Vibepay\FY 2025-26\01 Statements and Records\03 Subpaisa Payin Report\Transaction Report",
            'sabpaisa.settelment_report': r"\\server\F DRIVE DESKTOP\Parties Record\Vibepay\FY 2025-26\01 Statements and Records\03 Subpaisa Payin Report\Settlement Report",
            'rozarpayx.payout_report':    r"\\Server\F DRIVE DESKTOP\Parties Record\Swiftsend\2.F.Y. 2025-26\08 RazorPayX\Reports",
            'rozarpayx.transaction_report': r"\\Server\F DRIVE DESKTOP\Parties Record\Swiftsend\2.F.Y. 2025-26\08 RazorPayX\Reports",
            'swiftsend_portal.swiftsend_payout_report':   r"\\server\F DRIVE DESKTOP\Parties Record\Swiftsend\2.F.Y. 2025-26\Payout Report (Daily)",
            'swiftsend_portal.swiftsend_payout_wallet':   r"\\server\F DRIVE DESKTOP\Parties Record\Swiftsend\2.F.Y. 2025-26\Payout Wallet",
            'swiftsend_portal.swiftsend_finzen_payout':   r"\\server\F DRIVE DESKTOP\Parties Record\Swiftsend\2.F.Y. 2025-26\07 Finzen\Payout Statement",
            'swiftsend_portal.swiftsend_add_fund_report': r"\\server\F DRIVE DESKTOP\Parties Record\Swiftsend\2.F.Y. 2025-26\Add fund report",
            'payonetic_portal.payonetic_payout_report':   r"\\server\F DRIVE DESKTOP\Parties Record\Swiftsend\2.F.Y. 2025-26\05 Payonetic\Payout Report (Daily)",
            'payonetic_portal.payonetic_payout_wallet':   r"\\server\F DRIVE DESKTOP\Parties Record\Swiftsend\2.F.Y. 2025-26\05 Payonetic\Wallet",
            'payonetic_portal.payonetic_add_fund_report': r"\\server\F DRIVE DESKTOP\Parties Record\Swiftsend\2.F.Y. 2025-26\05 Payonetic\Add fund",
        }
        return locations.get(f"{db}.{table}", os.getcwd())

    def log_status(self, message):
        ts  = datetime.datetime.now().strftime("%H:%M:%S")
        msg = f"[{ts}] {message}\n"
        def upd():
            self.status_text.insert(tk.END, msg)
            self.status_text.see(tk.END)
            self.root.update_idletasks()
        if threading.current_thread() == threading.main_thread(): upd()
        else: self.root.after(0, upd)

    def clear_log(self):
        self.status_text.delete(1.0, tk.END)

    def check_password(self):
        import hashlib
        PASSWORD_HASH = "240be518fabd2724ddb6f04eeb1da5967448d7e831c08c8fa822809f74c720a9"
        MAX_ATTEMPTS  = 3
        self.attempts = 0

        dlg = tk.Toplevel(self.root)
        dlg.title("Secure Login")
        dlg.geometry("400x300")
        dlg.configure(bg="#f5f7fa")
        dlg.resizable(False, False)
        dlg.transient(self.root)
        dlg.grab_set(); dlg.focus_force(); dlg.lift()
        dlg.attributes('-topmost', True)
        dlg.update_idletasks()
        x = (dlg.winfo_screenwidth()-400)//2
        y = (dlg.winfo_screenheight()-300)//2
        dlg.geometry(f"+{x}+{y}")
        dlg.protocol("WM_DELETE_WINDOW", lambda: self.root.quit())

        mf = tk.Frame(dlg, bg="white", padx=25, pady=20)
        mf.pack(pady=20, padx=20, fill="both", expand=True)
        tk.Label(mf, text="Admin Authentication", font=("Segoe UI",14,"bold"),
                 bg="white", fg="#2c3e50").pack()
        tk.Label(mf, text="Enter your admin password to continue.",
                 font=("Segoe UI",10), bg="white", fg="#7f8c8d").pack(pady=(0,16))

        pv = tk.StringVar()
        pe = tk.Entry(mf, textvariable=pv, show="*", font=("Segoe UI",11))
        pe.pack(fill="x", pady=(0,8)); pe.focus_set()

        sl = tk.Label(mf, text="", fg="#e74c3c", font=("Segoe UI",9), bg="white")
        sl.pack()

        def verify():
            self.attempts += 1
            if hashlib.sha256(pv.get().encode()).hexdigest() == PASSWORD_HASH:
                dlg.destroy()
                self.root.lift(); self.root.focus_force()
            else:
                rem = MAX_ATTEMPTS - self.attempts
                if rem > 0:
                    sl.config(text=f"Incorrect. {rem} attempt(s) left")
                    pv.set(""); pe.focus_set()
                else:
                    messagebox.showerror("Access Denied","Max attempts reached.", parent=dlg)
                    self.root.quit()

        tk.Button(mf, text="Unlock System", font=("Segoe UI",10,"bold"),
                  bg="#3498db", fg="white", padx=10, pady=6, relief="flat",
                  command=verify).pack(fill="x", pady=(12,0))
        tk.Label(mf, text="For authorized personnel only",
                 font=("Segoe UI",8), fg="#bdc3c7", bg="white").pack(side="bottom", pady=(8,0))
        dlg.bind('<Return>', lambda e: verify())


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────
def main():
    root = tk.Tk()
    try: root.iconbitmap(default='icon.ico')
    except: pass
    app = MySQLExcelUploaderGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()